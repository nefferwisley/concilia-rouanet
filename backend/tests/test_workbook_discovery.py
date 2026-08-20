import io
import openpyxl
import pytest
from backend.services.workbook_discovery import discover_base_sheet


@pytest.fixture
def six_sheet_workbook():
    wb = openpyxl.Workbook()
    # Sheet 1: Capa
    ws1 = wb.active
    ws1.title = "Capa"
    ws1["A1"] = "MINISTÉRIO DA CULTURA"

    # Sheet 2: Conciliação Bancária
    ws2 = wb.create_sheet("Conciliação Bancária")
    ws2["A1"] = "PROJETO PRONAC 1961"
    ws2["A2"] = "RELATÓRIO DE PAGAMENTOS"
    ws2["A3"] = "Item"
    ws2["B3"] = "Data Pagamento"
    ws2["C3"] = "Fornecedor / Favorecido"
    ws2["D3"] = "Documento Fiscal"
    ws2["E3"] = "Rubrica Orçamentária"
    ws2["F3"] = "Valor Pago (R$)"
    ws2["A4"] = 1
    ws2["B4"] = "2024-01-15"
    ws2["C4"] = "Fornecedor Alpha LTDA"
    ws2["D4"] = "NF 1234"
    ws2["E4"] = "Direção e Produção"
    ws2["F4"] = 1500.50

    # Sheets 3-6: Outras abas
    wb.create_sheet("Rubricas")
    wb.create_sheet("Extrato")
    wb.create_sheet("Controle Interno")
    wb.create_sheet("Resumo")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_discovers_conciliacao_sheet_in_mislabeled_workbook(six_sheet_workbook):
    candidate = discover_base_sheet(six_sheet_workbook)
    assert candidate is not None
    assert "concilia" in candidate.sheet_name.casefold()
    assert candidate.header_row == 3
    assert candidate.column_map.get("fornecedor") == "C"
    assert candidate.column_map.get("valor") == "F"
