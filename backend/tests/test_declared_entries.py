from decimal import Decimal
import io
import openpyxl
import pytest
from backend.services.declared_entries import parse_declared_entries
from backend.services.workbook_discovery import SheetCandidate


def test_parses_declared_entries_with_cell_provenance():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conciliacao"
    ws["A1"] = "Item"
    ws["B1"] = "Data"
    ws["C1"] = "Fornecedor"
    ws["D1"] = "Valor"

    ws["A2"] = 1
    ws["B2"] = "15/05/2024"
    ws["C2"] = "Empresa Cultural ME"
    ws["D2"] = 2500.75

    ws["A3"] = "Total"
    ws["D3"] = 2500.75

    buf = io.BytesIO()
    wb.save(buf)
    content = buf.getvalue()

    candidate = SheetCandidate(
        sheet_name="Conciliacao",
        header_row=1,
        column_map={"controle": "A", "data": "B", "fornecedor": "C", "valor": "D"},
        confidence=0.9,
    )

    entries = parse_declared_entries(content, candidate)

    assert len(entries) == 1
    entry = entries[0]
    assert entry.row_number == 2
    assert entry.fornecedor_declarado == "Empresa Cultural ME"
    assert entry.valor_declarado == Decimal("2500.75")
    assert entry.cell_locators.get("valor") == "D2"
    assert entry.cell_locators.get("fornecedor") == "C2"
