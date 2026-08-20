import io
import unicodedata
from dataclasses import dataclass
from typing import Optional

import openpyxl


@dataclass
class SheetCandidate:
    sheet_name: str
    header_row: int
    column_map: dict[str, str]
    confidence: float


def _normalize(text: str) -> str:
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(text))
    return nfkd.encode("ascii", "ignore").decode("ascii").lower().strip()


def _score_header_row(row_values: list[tuple[str, str]]) -> tuple[float, dict[str, str]]:
    """
    row_values: list of (col_letter, cell_value_str)
    """
    mapping: dict[str, str] = {}
    score = 0.0

    keywords = {
        "controle": ["item", "controle", "no", "num", "sequencia", "ordem"],
        "data": ["data", "dt", "dt.", "pagamento", "vencimento", "emissao"],
        "fornecedor": ["fornecedor", "favorecido", "beneficiario", "credor", "nome", "razao"],
        "documento": ["documento", "doc", "nf", "nota", "recibo", "comprovante", "fatura"],
        "rubrica": ["rubrica", "etapa", "meta", "item orcamentario", "orcamento", "categoria"],
        "valor": ["valor", "vlr", "pago", "bruto", "liquido", "total", "r$"],
    }

    for col_letter, raw_val in row_values:
        norm = _normalize(raw_val)
        if not norm:
            continue

        for role, terms in keywords.items():
            if role not in mapping:
                if any(term in norm for term in terms):
                    mapping[role] = col_letter
                    if role in ("fornecedor", "valor", "data"):
                        score += 0.25
                    else:
                        score += 0.1

    return score, mapping


def discover_base_sheet(content: bytes) -> Optional[SheetCandidate]:
    if not content:
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception:
        return None

    best_candidate: Optional[SheetCandidate] = None
    best_score = 0.0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_norm = _normalize(sheet_name)

        bonus = 0.0
        if "concilia" in name_norm or "pagamento" in name_norm or "despesa" in name_norm or "relatorio" in name_norm:
            bonus = 0.2

        # Inspeciona até 20 primeiras linhas
        row_idx = 0
        for row in ws.iter_rows(values_only=False):
            row_idx += 1
            if row_idx > 20:
                break

            cells: list[tuple[str, str]] = []
            for cell in row:
                if cell.value is not None:
                    col_letter = openpyxl.utils.get_column_letter(cell.column)
                    cells.append((col_letter, str(cell.value)))

            if len(cells) < 2:
                continue

            score, col_map = _score_header_row(cells)
            total_score = score + bonus

            if total_score > best_score and ("valor" in col_map or "fornecedor" in col_map):
                best_score = total_score
                best_candidate = SheetCandidate(
                    sheet_name=sheet_name,
                    header_row=row_idx,
                    column_map=col_map,
                    confidence=min(1.0, total_score),
                )

    return best_candidate
