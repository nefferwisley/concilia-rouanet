from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
import io
import re
from typing import Any, Optional

import openpyxl

from backend.services.workbook_discovery import SheetCandidate


@dataclass
class DeclaredEntry:
    row_number: int
    valor_declarado: Optional[Decimal] = None
    data_declarada: Optional[str] = None
    fornecedor_declarado: Optional[str] = None
    rubrica_declarada: Optional[str] = None
    documento_declarado: Optional[str] = None
    descricao_declarada: Optional[str] = None
    cell_locators: dict[str, str] = field(default_factory=dict)
    raw_values: dict[str, Any] = field(default_factory=dict)


def _parse_decimal(val: Any) -> Optional[Decimal]:
    if val is None:
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    text = str(val).strip()
    if not text:
        return None
    # Remove R$, espaços
    text = re.sub(r"[^\d,\.-]", "", text)
    # Se padrão brasileiro 1.234,56
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except Exception:
        return None


def _parse_date(val: Any) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    text = str(val).strip()
    if not text:
        return None
    # Detecta DD/MM/YYYY
    match = re.match(r"^(\d{1,2})[/.-](\d{1,2})[/.-](\d{2,4})", text)
    if match:
        d, m, y = match.groups()
        if len(y) == 2:
            y = f"20{y}"
        return f"{y.zfill(4)}-{m.zfill(2)}-{d.zfill(2)}"
    return text


def parse_declared_entries(content: bytes, candidate: SheetCandidate) -> list[DeclaredEntry]:
    if not content or not candidate:
        return []

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    if candidate.sheet_name not in wb.sheetnames:
        return []

    ws = wb[candidate.sheet_name]
    col_map = candidate.column_map

    entries: list[DeclaredEntry] = []

    for row_idx in range(candidate.header_row + 1, ws.max_row + 1):
        row_cells = {
            openpyxl.utils.get_column_letter(cell.column): cell
            for cell in ws[row_idx]
            if cell.value is not None
        }

        if not row_cells:
            continue

        # Detecta se é linha de total / subtotal
        row_texts = [str(c.value).lower() for c in row_cells.values()]
        if any("total" in t or "subtotal" in t or "saldo" in t for t in row_texts):
            continue

        cell_locators: dict[str, str] = {}
        raw_values: dict[str, Any] = {}

        valor_val = None
        if "valor" in col_map and col_map["valor"] in row_cells:
            col_letter = col_map["valor"]
            c = row_cells[col_letter]
            valor_val = _parse_decimal(c.value)
            cell_locators["valor"] = f"{col_letter}{row_idx}"
            raw_values["valor"] = str(c.value)

        fornecedor_val = None
        if "fornecedor" in col_map and col_map["fornecedor"] in row_cells:
            col_letter = col_map["fornecedor"]
            c = row_cells[col_letter]
            fornecedor_val = str(c.value).strip() if c.value is not None else None
            cell_locators["fornecedor"] = f"{col_letter}{row_idx}"
            raw_values["fornecedor"] = str(c.value)

        data_val = None
        if "data" in col_map and col_map["data"] in row_cells:
            col_letter = col_map["data"]
            c = row_cells[col_letter]
            data_val = _parse_date(c.value)
            cell_locators["data"] = f"{col_letter}{row_idx}"
            raw_values["data"] = str(c.value)

        rubrica_val = None
        if "rubrica" in col_map and col_map["rubrica"] in row_cells:
            col_letter = col_map["rubrica"]
            c = row_cells[col_letter]
            rubrica_val = str(c.value).strip() if c.value is not None else None
            cell_locators["rubrica"] = f"{col_letter}{row_idx}"
            raw_values["rubrica"] = str(c.value)

        doc_val = None
        if "documento" in col_map and col_map["documento"] in row_cells:
            col_letter = col_map["documento"]
            c = row_cells[col_letter]
            doc_val = str(c.value).strip() if c.value is not None else None
            cell_locators["documento"] = f"{col_letter}{row_idx}"
            raw_values["documento"] = str(c.value)

        # Se a linha não tem nem valor nem fornecedor, ignora
        if valor_val is None and not fornecedor_val:
            continue

        entries.append(
            DeclaredEntry(
                row_number=row_idx,
                valor_declarado=valor_val,
                data_declarada=data_val,
                fornecedor_declarado=fornecedor_val,
                rubrica_declarada=rubrica_val,
                documento_declarado=doc_val,
                cell_locators=cell_locators,
                raw_values=raw_values,
            )
        )

    return entries
